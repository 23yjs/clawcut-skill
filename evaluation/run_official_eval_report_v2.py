from __future__ import annotations

import argparse
import csv
import html
import json
import os
import time
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

try:
    from .ark_aesthetic_judge_client import ArkAestheticJudgeConfig
    from .ark_resolver_client import ArkResolverConfig
    from .auto_eval import AutoEvalConfig, run_auto_eval
    from .dover_quality import build_dover_config
    from .issue_summary import ISSUE_LABELS, build_issue_summary
    from .tos_uploader import build_tos_upload_config
except ImportError:  # pragma: no cover - script mode
    from ark_aesthetic_judge_client import ArkAestheticJudgeConfig
    from ark_resolver_client import ArkResolverConfig
    from auto_eval import AutoEvalConfig, run_auto_eval
    from dover_quality import build_dover_config
    from issue_summary import ISSUE_LABELS, build_issue_summary
    from tos_uploader import build_tos_upload_config


SUMMARY_SCHEMA_VERSION = "official_eval_summary_v2"
EVALUATION_RESULT_SCHEMA_VERSION = "evaluation_result_v2"
XLSX_SHEETS = [
    "全部 Case 结果",
    "内容选择明细",
    "成片体验明细",
    "技术质量明细",
    "耗时与 Token 明细",
    "指标说明",
]

FAILURE_STATUS_LABELS = {
    "invalid_artifact": "产物校验失败",
    "judge_failed": "成片观看体验评测失败",
    "judge_video_upload_failed": "Judge 视频上传失败",
    "resolver_failed": "用户指令解析失败",
    "missing_evaluation_result": "缺少评测结果",
    "batch_case_failed": "单条 Case 执行失败",
}

USER_REQUIREMENT_LABELS = {
    "generic_selection": "没有明确要求，由 Skill 自主提炼高光",
    "specific_preferential": "重点突出用户指定的内容",
    "specific_exclusive": "只保留用户指定的内容",
    "conflict_preferential": "保留重点，同时排除指定内容",
    "conflict_exclusive": "严格保留目标，并排除指定内容",
    "duration_control": "按照目标时长取舍内容",
}

TEST_TYPE_TO_USER_REQUIREMENT = {
    "baseline_generic": "generic_selection",
    "specific_following": "specific_preferential",
    "conflict_exclusion": "conflict_preferential",
    "duration_constrained": "duration_control",
}

VIDEO_SCENARIO_LABELS = {
    "sparse_highlight": "高光片段较少的视频",
    "dense_highlight": "高光片段较密集的视频",
    "high_dynamic": "动作变化较快的视频",
    "long_video": "长视频",
    "multi_topic": "包含多个主题的视频",
    "speech_dense": "口播内容较密集的视频",
    "repetitive_process": "重复步骤较多的视频",
    "out_of_prompt_type": "未预设类型的视频",
    "replay_dedup": "包含重复回放的视频",
}

IGNORED_CHALLENGE_TAGS = {"", "按视频自身类型继承", "none", "null"}


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            text = line.strip()
            if text:
                rows.append(json.loads(text))
    return rows


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_path_maps(values: list[str] | None) -> dict[str, str]:
    path_map: dict[str, str] = {}
    for value in values or []:
        if "=" not in value:
            raise ValueError(f"--path-map must use FROM=TO format: {value}")
        source, target = value.split("=", 1)
        path_map[source.strip().rstrip("/")] = target.strip().rstrip("/")
    return path_map


def map_path(value: Any, path_map: dict[str, str] | None = None) -> Path:
    text = str(value)
    for source_prefix, target_prefix in (path_map or {}).items():
        if text == source_prefix:
            return Path(target_prefix)
        if text.startswith(source_prefix + "/"):
            return Path(target_prefix + text[len(source_prefix) :])
    return Path(text)


def sum_numbers(*values: Any) -> int | float | None:
    numbers = [value for value in values if isinstance(value, (int, float))]
    return sum(numbers) if numbers else None


def result_summary_from(result: dict[str, Any]) -> dict[str, Any]:
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    return artifact.get("result_summary") if isinstance(artifact.get("result_summary"), dict) else {}


def resolver_usage_from(result: dict[str, Any]) -> dict[str, Any]:
    metadata = result.get("resolver_metadata") if isinstance(result.get("resolver_metadata"), dict) else {}
    return metadata.get("resolver_usage") if isinstance(metadata.get("resolver_usage"), dict) else {}


def judge_usage_from(result: dict[str, Any]) -> dict[str, Any]:
    aesthetic = result.get("aesthetic_judge") if isinstance(result.get("aesthetic_judge"), dict) else {}
    metadata = aesthetic.get("judge_metadata")
    if not isinstance(metadata, list):
        metadata = []
    usage_items = [
        item.get("aesthetic_judge_usage")
        for item in metadata
        if isinstance(item, dict) and isinstance(item.get("aesthetic_judge_usage"), dict)
    ]
    latency = sum_numbers(*[
        item.get("aesthetic_judge_latency_seconds")
        for item in metadata
        if isinstance(item, dict)
    ])
    return {
        "latency_seconds": round(float(latency), 3) if latency is not None else None,
        "prompt_tokens": sum_numbers(*[usage.get("prompt_tokens") for usage in usage_items]),
        "completion_tokens": sum_numbers(*[usage.get("completion_tokens") for usage in usage_items]),
        "total_tokens": sum_numbers(*[usage.get("total_tokens") for usage in usage_items]),
    }


def build_consumption(result: dict[str, Any]) -> dict[str, Any]:
    summary = result_summary_from(result)
    skill_consumption = summary.get("skill_consumption") if isinstance(summary.get("skill_consumption"), dict) else {}
    resolver_metadata = result.get("resolver_metadata") if isinstance(result.get("resolver_metadata"), dict) else {}
    resolver_usage = resolver_usage_from(result)
    judge_usage = judge_usage_from(result)
    skill_total = skill_consumption.get("skill_llm_total_tokens", summary.get("skill_llm_total_tokens"))
    resolver_total = resolver_usage.get("total_tokens")
    judge_total = judge_usage.get("total_tokens")
    evaluation_total = sum_numbers(resolver_total, judge_total)
    end_to_end_total = sum_numbers(skill_total, resolver_total, judge_total)
    skill_elapsed = skill_consumption.get("skill_run_elapsed_seconds", summary.get("skill_run_elapsed_seconds"))
    evaluation_elapsed = result.get("evaluation_elapsed_seconds")
    end_to_end_elapsed = (
        skill_elapsed + evaluation_elapsed
        if isinstance(skill_elapsed, (int, float)) and isinstance(evaluation_elapsed, (int, float))
        else None
    )
    return {
        "video_editing": {
            "elapsed_seconds": skill_elapsed,
            "preview_generation_seconds": skill_consumption.get("preview_generation_seconds", summary.get("preview_generation_seconds")),
            "skill_llm_latency_seconds": skill_consumption.get("skill_llm_latency_seconds", summary.get("skill_llm_latency_seconds")),
            "ffmpeg_render_seconds": skill_consumption.get("ffmpeg_render_seconds", summary.get("ffmpeg_render_seconds")),
            "llm_prompt_tokens": skill_consumption.get("skill_llm_prompt_tokens", summary.get("skill_llm_prompt_tokens")),
            "llm_completion_tokens": skill_consumption.get("skill_llm_completion_tokens", summary.get("skill_llm_completion_tokens")),
            "llm_total_tokens": skill_total,
            "skill_llm_attempt_count": skill_consumption.get("skill_llm_attempt_count", summary.get("skill_llm_attempt_count")),
        },
        "evaluation": {
            "elapsed_seconds": evaluation_elapsed,
            "resolver_latency_seconds": resolver_metadata.get("resolver_latency_seconds"),
            "resolver_prompt_tokens": resolver_usage.get("prompt_tokens"),
            "resolver_completion_tokens": resolver_usage.get("completion_tokens"),
            "resolver_total_tokens": resolver_total,
            "judge_latency_seconds": judge_usage.get("latency_seconds"),
            "judge_prompt_tokens": judge_usage.get("prompt_tokens"),
            "judge_completion_tokens": judge_usage.get("completion_tokens"),
            "judge_total_tokens": judge_total,
            "llm_total_tokens": evaluation_total,
        },
        "end_to_end": {
            "elapsed_seconds": end_to_end_elapsed,
            "llm_total_tokens": end_to_end_total,
        },
    }


def build_case_metadata(case: dict[str, Any]) -> dict[str, Any]:
    keys = [
        "test_type",
        "primary_capability",
        "tested_capability",
        "challenge_tags",
        "execution_tier",
        "include_in_official_score",
        "priority",
    ]
    return {key: case.get(key) for key in keys if key in case}


def official_eligibility(result: dict[str, Any]) -> dict[str, Any]:
    reasons: list[str] = []
    if result.get("evaluation_status") != "scored_complete":
        reasons.append(str(result.get("evaluation_status") or "not_scored_complete"))
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    if artifact.get("fallback_used"):
        reasons.append("skill_fallback_used")
    if artifact.get("skill_backend_used") not in {None, "ark"}:
        reasons.append("skill_backend_not_ark")
    if artifact and not artifact.get("artifact_validation_passed"):
        reasons.append("artifact_validation_failed")
    technical = result.get("technical_quality") if isinstance(result.get("technical_quality"), dict) else {}
    if technical and technical.get("technical_quality_passed") is False:
        reasons.append("technical_quality_failed")
    return {"eligible": not reasons, "exclusion_reasons": reasons}


def enrich_result_v2(case: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(result)
    enriched["evaluation_result_schema_version"] = EVALUATION_RESULT_SCHEMA_VERSION
    enriched["case_id"] = case.get("case_id")
    enriched["case_metadata"] = build_case_metadata(case)
    enriched["official_score_eligibility"] = official_eligibility(enriched)
    enriched["failure_info"] = {
        "error_type": enriched.get("error_type"),
        "error_message": enriched.get("error_message"),
    }
    artifact = enriched.get("artifact_validation") if isinstance(enriched.get("artifact_validation"), dict) else {}
    enriched["fallback_info"] = {
        "fallback_used": artifact.get("fallback_used"),
        "skill_backend_used": artifact.get("skill_backend_used"),
    }
    enriched["manual_review"] = {
        "required": enriched.get("evaluation_status") in {"manual_review_required", "judge_manual_review_required"},
        "recommended": bool((enriched.get("editing_experience") or {}).get("manual_review_recommended")) if isinstance(enriched.get("editing_experience"), dict) else False,
    }
    enriched["issue_summary"] = build_issue_summary(enriched)
    enriched["consumption"] = build_consumption(enriched)
    return enriched


def run_case(
    *,
    case: dict[str, Any],
    index: int,
    args: argparse.Namespace,
    path_map: dict[str, str],
) -> dict[str, Any]:
    case_id = str(case.get("case_id") or f"case_{index:03d}")
    run_dir = args.output_dir / "runs" / case_id
    result_path = run_dir / "evaluation_result.json"
    if args.resume and result_path.exists() and not args.retry_failed:
        return json.loads(result_path.read_text(encoding="utf-8"))
    if args.retry_failed and result_path.exists():
        existing = json.loads(result_path.read_text(encoding="utf-8"))
        if existing.get("official_score_eligibility", {}).get("eligible"):
            return existing
    result = run_auto_eval(
        AutoEvalConfig(
            input_video=map_path(case["input_video"], path_map),
            instruction=str(case["instruction"]),
            target_duration=case.get("target_duration"),
            skill_output_dir=map_path(case["skill_output_dir"], path_map),
            gt_dir=args.gt_dir,
            output_dir=run_dir,
            resolver_config=ArkResolverConfig(),
            judge_video_url=case.get("judge_video_url"),
            aesthetic_judge_config=ArkAestheticJudgeConfig(),
            dover_config=build_dover_config(enabled=False),
            technical_quality_config=args.technical_quality_config,
            auto_upload_judge_video=bool(args.auto_upload_judge_video),
            tos_upload_config=build_tos_upload_config(
                enabled=bool(args.auto_upload_judge_video),
                bucket=args.tos_bucket,
                region=args.tos_region,
                endpoint=args.tos_endpoint,
                key_prefix=args.tos_key_prefix,
                presign_expires_seconds=args.tos_presign_expires_seconds,
            ),
            path_map=path_map,
            eval_run_id=args.output_dir.name,
            case_id=case_id,
            skill_run_id=str(case.get("skill_run_id") or Path(str(case.get("skill_output_dir") or "")).name or "run_01"),
        )
    )
    enriched = enrich_result_v2(case, result)
    write_json(result_path, enriched)
    return enriched


def load_existing_result(case: dict[str, Any], index: int, output_dir: Path) -> dict[str, Any]:
    case_id = str(case.get("case_id") or f"case_{index:03d}")
    path = output_dir / "runs" / case_id / "evaluation_result.json"
    if not path.exists():
        return enrich_result_v2(case, {
            "evaluation_status": "missing_evaluation_result",
            "evaluation_scope": "failed",
            "selection_score_v1": None,
            "editing_experience_score_v1": None,
            "aesthetic_score_v1": None,
            "final_score_v2": None,
            "error_type": "MissingEvaluationResult",
            "error_message": f"evaluation_result.json not found: {path}",
        })
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if payload.get("evaluation_result_schema_version") == EVALUATION_RESULT_SCHEMA_VERSION else enrich_result_v2(case, payload)


def flat_case_row(case: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    consumption = result.get("consumption") if isinstance(result.get("consumption"), dict) else {}
    editing = consumption.get("video_editing", {}) if isinstance(consumption.get("video_editing"), dict) else {}
    evaluation = consumption.get("evaluation", {}) if isinstance(consumption.get("evaluation"), dict) else {}
    e2e = consumption.get("end_to_end", {}) if isinstance(consumption.get("end_to_end"), dict) else {}
    technical = result.get("technical_quality") if isinstance(result.get("technical_quality"), dict) else {}
    issues = result.get("issue_summary") if isinstance(result.get("issue_summary"), list) else []
    return {
        "case_id": result.get("case_id") or case.get("case_id"),
        "video_id": result.get("video_id") or case.get("video_id"),
        "test_type": case.get("test_type"),
        "instruction": case.get("instruction"),
        "evaluation_status": result.get("evaluation_status"),
        "official_eligible": (result.get("official_score_eligibility") or {}).get("eligible"),
        "selection_score_v1": result.get("selection_score_v1"),
        "editing_experience_score_v1": result.get("editing_experience_score_v1"),
        "final_score_v2": result.get("final_score_v2"),
        "technical_quality_passed": technical.get("technical_quality_passed"),
        "issues": "；".join(issue.get("label", "") for issue in issues),
        "video_editing_elapsed_seconds": editing.get("elapsed_seconds"),
        "preview_generation_seconds": editing.get("preview_generation_seconds"),
        "skill_llm_latency_seconds": editing.get("skill_llm_latency_seconds"),
        "ffmpeg_render_seconds": editing.get("ffmpeg_render_seconds"),
        "skill_llm_prompt_tokens": editing.get("llm_prompt_tokens"),
        "skill_llm_completion_tokens": editing.get("llm_completion_tokens"),
        "skill_llm_total_tokens": editing.get("llm_total_tokens"),
        "evaluation_elapsed_seconds": evaluation.get("elapsed_seconds"),
        "resolver_latency_seconds": evaluation.get("resolver_latency_seconds"),
        "resolver_total_tokens": evaluation.get("resolver_total_tokens"),
        "judge_latency_seconds": evaluation.get("judge_latency_seconds"),
        "judge_total_tokens": evaluation.get("judge_total_tokens"),
        "end_to_end_elapsed_seconds": e2e.get("elapsed_seconds"),
        "end_to_end_llm_total_tokens": e2e.get("llm_total_tokens"),
    }


def average(values: list[Any]) -> float | None:
    numbers = [float(value) for value in values if isinstance(value, (int, float))]
    return round(sum(numbers) / len(numbers), 3) if numbers else None


def total(values: list[Any]) -> float | None:
    numbers = [float(value) for value in values if isinstance(value, (int, float))]
    return round(sum(numbers), 3) if numbers else None


def build_summary(cases: list[dict[str, Any]], results: list[dict[str, Any]], rows: list[dict[str, Any]], output_dir: Path) -> dict[str, Any]:
    eligible = [result for result in results if (result.get("official_score_eligibility") or {}).get("eligible")]
    status_counts: dict[str, int] = {}
    categories: list[str] = []
    for result in results:
        status = str(result.get("evaluation_status") or "unknown")
        status_counts[status] = status_counts.get(status, 0) + 1
        categories.append(_result_category(result))
    result_distribution = {
        "official_scored": categories.count("official_scored"),
        "fallback_diagnostic": categories.count("fallback_diagnostic"),
        "manual_review_blocked": categories.count("manual_review_blocked"),
        "failed_or_incomplete": categories.count("failed_or_incomplete"),
    }
    unclassified_count = sum(1 for category in categories if category not in result_distribution)
    assert sum(result_distribution.values()) == len(cases)
    failure_summary = _failure_summary(results, categories)
    fallback_summary = _fallback_summary(results)
    manual_review_summary = _manual_review_summary(results)
    return {
        "summary_schema_version": SUMMARY_SCHEMA_VERSION,
        "run_metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "case_count": len(cases),
            "output_dir": str(output_dir),
        },
        "execution_overview": {
            "planned_case_count": len(cases),
            "generated_result_count": sum(1 for result in results if _has_valid_generated_artifact(result)),
            "official_scored_count": len(eligible),
            "fallback_diagnostic_count": result_distribution["fallback_diagnostic"],
            "manual_review_blocked_count": result_distribution["manual_review_blocked"],
            "failed_or_incomplete_count": result_distribution["failed_or_incomplete"],
            "unclassified_count": unclassified_count,
            "status_counts": status_counts,
            "result_distribution": result_distribution,
        },
        "overall_scores": {
            "selection_score_v1_avg": average([item.get("selection_score_v1") for item in eligible]),
            "editing_experience_score_v1_avg": average([item.get("editing_experience_score_v1") for item in eligible]),
            "final_score_v2_avg": average([item.get("final_score_v2") for item in eligible]),
        },
        "consumption_overview": {
            "video_editing_elapsed_total": total([row.get("video_editing_elapsed_seconds") for row in rows]),
            "video_editing_elapsed_avg": average([row.get("video_editing_elapsed_seconds") for row in rows]),
            "skill_llm_total_tokens_total": total([row.get("skill_llm_total_tokens") for row in rows]),
            "evaluation_elapsed_total": total([row.get("evaluation_elapsed_seconds") for row in rows]),
            "evaluation_elapsed_avg": average([row.get("evaluation_elapsed_seconds") for row in rows]),
            "resolver_total_tokens_total": total([row.get("resolver_total_tokens") for row in rows]),
            "judge_total_tokens_total": total([row.get("judge_total_tokens") for row in rows]),
            "end_to_end_elapsed_total": total([row.get("end_to_end_elapsed_seconds") for row in rows]),
            "end_to_end_llm_total_tokens_total": total([row.get("end_to_end_llm_total_tokens") for row in rows]),
        },
        "failure_summary": failure_summary,
        "fallback_summary": fallback_summary,
        "manual_review_summary": manual_review_summary,
        "breakdowns": {
            "by_issue": _issue_breakdown(results),
            "by_status": status_counts,
            "by_user_requirement": _user_requirement_breakdown(cases, results),
            "by_video_scenario": _video_scenario_breakdown(cases, results),
            "by_test_type": _breakdown_by(rows, "test_type"),
        },
        "case_index": rows,
        "special_reports_overview": {},
        "artifact_paths": {
            "case_results_csv": str(output_dir / "case_results.csv"),
            "case_results_xlsx": str(output_dir / "case_results.xlsx"),
            "report_html": str(output_dir / "report.html"),
            "cases_dir": str(output_dir / "cases"),
        },
    }


def _result_category(result: dict[str, Any]) -> str:
    if (result.get("official_score_eligibility") or {}).get("eligible"):
        return "official_scored"
    if (result.get("fallback_info") or {}).get("fallback_used"):
        return "fallback_diagnostic"
    if (result.get("manual_review") or {}).get("required"):
        return "manual_review_blocked"
    return "failed_or_incomplete"


def _case_id(result: dict[str, Any]) -> str:
    return str(result.get("case_id") or result.get("video_id") or "")


def _has_valid_generated_artifact(result: dict[str, Any]) -> bool:
    artifact = result.get("artifact_validation")
    return isinstance(artifact, dict) and artifact.get("artifact_validation_passed") is True


def _failure_summary(results: list[dict[str, Any]], categories: list[str]) -> dict[str, Any]:
    failed_results = [
        result
        for result, category in zip(results, categories)
        if category == "failed_or_incomplete"
    ]
    case_ids = [_case_id(result) for result in failed_results]
    by_status_map: dict[str, list[str]] = {}
    for result in failed_results:
        status = str(result.get("evaluation_status") or "unknown")
        by_status_map.setdefault(status, []).append(_case_id(result))
    by_status = [
        {
            "status": status,
            "label": FAILURE_STATUS_LABELS.get(status, status),
            "count": len(ids),
            "case_ids": ids,
        }
        for status, ids in sorted(by_status_map.items())
    ]
    return {
        "count": len(failed_results),
        "failed_count": len(failed_results),
        "by_status": by_status,
        "case_ids": case_ids,
    }


def _fallback_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    case_ids = [
        _case_id(result)
        for result in results
        if (result.get("fallback_info") or {}).get("fallback_used")
    ]
    return {
        "count": len(case_ids),
        "skill_fallback_count": len(case_ids),
        "label": "Ark 调用失败，已触发兜底",
        "case_ids": case_ids,
    }


def _manual_review_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    required_case_ids = [
        _case_id(result)
        for result in results
        if (result.get("manual_review") or {}).get("required")
    ]
    recommended_case_ids = [
        _case_id(result)
        for result in results
        if (result.get("manual_review") or {}).get("recommended")
    ]
    return {
        "required_count": len(required_case_ids),
        "recommended_count": len(recommended_case_ids),
        "required_case_ids": required_case_ids,
        "recommended_case_ids": recommended_case_ids,
    }


def _issue_breakdown(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = {}
    for result in results:
        for issue in result.get("issue_summary", []) or []:
            key = str(issue.get("issue_type") or "")
            if not key:
                continue
            bucket = buckets.setdefault(
                key,
                {
                    "key": key,
                    "label": str(issue.get("label") or ISSUE_LABELS.get(key, key)),
                    "count": 0,
                    "case_ids": [],
                },
            )
            bucket["count"] += 1
            bucket["case_ids"].append(_case_id(result))
    return sorted(buckets.values(), key=lambda item: (-int(item["count"]), str(item["key"])))


def _user_requirement_key(case: dict[str, Any]) -> str:
    primary = str(case.get("primary_capability") or "").strip()
    if primary:
        return primary
    return TEST_TYPE_TO_USER_REQUIREMENT.get(str(case.get("test_type") or ""), str(case.get("test_type") or "unknown"))


def _user_requirement_breakdown(cases: list[dict[str, Any]], results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = {}
    for case, result in zip(cases, results):
        key = _user_requirement_key(case)
        _add_breakdown_case(
            buckets,
            key=key,
            label=USER_REQUIREMENT_LABELS.get(key, key),
            result=result,
        )
    return _finalize_breakdown_buckets(buckets)


def _challenge_tags(case: dict[str, Any]) -> list[str]:
    raw = case.get("challenge_tags")
    if isinstance(raw, list):
        values = raw
    elif isinstance(raw, str):
        values = raw.split(",")
    else:
        values = []
    tags = []
    for value in values:
        tag = str(value).strip()
        if tag and tag not in IGNORED_CHALLENGE_TAGS:
            tags.append(tag)
    return tags


def _video_scenario_breakdown(cases: list[dict[str, Any]], results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = {}
    for case, result in zip(cases, results):
        for key in _challenge_tags(case):
            _add_breakdown_case(
                buckets,
                key=key,
                label=VIDEO_SCENARIO_LABELS.get(key, key),
                result=result,
            )
    return _finalize_breakdown_buckets(buckets)


def _add_breakdown_case(
    buckets: dict[str, dict[str, Any]],
    *,
    key: str,
    label: str,
    result: dict[str, Any],
) -> None:
    bucket = buckets.setdefault(
        key,
        {
            "key": key,
            "label": label,
            "case_count": 0,
            "official_scored_count": 0,
            "average_content_selection_score": None,
            "average_editing_experience_score": None,
            "average_final_score": None,
            "case_ids": [],
            "_selection_scores": [],
            "_editing_scores": [],
            "_final_scores": [],
        },
    )
    bucket["case_count"] += 1
    bucket["case_ids"].append(_case_id(result))
    if (result.get("official_score_eligibility") or {}).get("eligible"):
        bucket["official_scored_count"] += 1
        bucket["_selection_scores"].append(result.get("selection_score_v1"))
        bucket["_editing_scores"].append(result.get("editing_experience_score_v1"))
        bucket["_final_scores"].append(result.get("final_score_v2"))


def _finalize_breakdown_buckets(buckets: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for bucket in buckets.values():
        row = dict(bucket)
        row["average_content_selection_score"] = average(row.pop("_selection_scores"))
        row["average_editing_experience_score"] = average(row.pop("_editing_scores"))
        row["average_final_score"] = average(row.pop("_final_scores"))
        rows.append(row)
    return sorted(rows, key=lambda item: (-int(item["case_count"]), str(item["key"])))


def _breakdown_by(rows: list[dict[str, Any]], field: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for row in rows:
        key = str(row.get(field) or "unknown")
        bucket = result.setdefault(key, {"case_count": 0, "eligible_count": 0, "final_score_v2_avg": None, "_scores": []})
        bucket["case_count"] += 1
        if row.get("official_eligible"):
            bucket["eligible_count"] += 1
            bucket["_scores"].append(row.get("final_score_v2"))
    for bucket in result.values():
        bucket["final_score_v2_avg"] = average(bucket.pop("_scores"))
    return result


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(rows[0].keys()) if rows else ["case_id"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    results = [_load_xlsx_case_result(path.parent, row) for row in rows]
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_payloads = [
        ("全部 Case 结果", _xlsx_all_case_rows(path.parent, rows, results)),
        ("内容选择明细", _xlsx_selection_rows(results)),
        ("成片体验明细", _xlsx_editing_rows(results)),
        ("技术质量明细", _xlsx_technical_rows(results)),
        ("耗时与 Token 明细", _xlsx_consumption_rows(results)),
        ("指标说明", _xlsx_metric_description_rows()),
    ]
    header_fill = PatternFill("solid", fgColor="12365F")
    header_font = Font(color="FFFFFF", bold=True)
    thin_fill_red = PatternFill("solid", fgColor="FDE9E7")
    thin_fill_orange = PatternFill("solid", fgColor="FCEFD8")
    thin_fill_blue = PatternFill("solid", fgColor="EAF2FF")
    thin_fill_yellow = PatternFill("solid", fgColor="FFF7D6")
    for sheet_name, payload in sheet_payloads:
        worksheet = workbook.create_sheet(sheet_name)
        headers = payload["headers"]
        worksheet.append(headers)
        for record in payload["rows"]:
            worksheet.append([record.get(header) for header in headers])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in worksheet.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, header in enumerate(headers, start=1):
            width = _xlsx_column_width(header, worksheet, index)
            worksheet.column_dimensions[get_column_letter(index)].width = width
        for column in payload.get("percent_columns", []):
            if column in headers:
                col_idx = headers.index(column) + 1
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for item in cell:
                        if isinstance(item.value, (int, float)):
                            item.number_format = "0.00%"
        for column in payload.get("score_columns", []):
            if column in headers:
                col_idx = headers.index(column) + 1
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for item in cell:
                        if isinstance(item.value, (int, float)):
                            item.number_format = "0.00"
        if sheet_name == "全部 Case 结果" and worksheet.max_row >= 2:
            data_range = f"A2:L{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(formula=['OR($E2="产物校验失败",$E2="成片观看体验评测失败",$E2="Judge 视频上传失败",$E2="用户指令解析失败",$E2="缺少评测结果",$E2="单条 Case 执行失败")'], fill=thin_fill_red),
            )
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(formula=['$I2="是"'], fill=thin_fill_orange),
            )
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(formula=['$J2="是"'], fill=thin_fill_blue),
            )
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(formula=["AND(ISNUMBER($H2),$H2<70)"], fill=thin_fill_yellow),
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


STATUS_LABELS = {
    "scored_complete": "正式计分完成",
    "diagnostic_only": "仅用于诊断",
    "invalid_artifact": "产物校验失败",
    "judge_failed": "成片观看体验评测失败",
    "judge_video_upload_failed": "Judge 视频上传失败",
    "resolver_failed": "用户指令解析失败",
    "missing_evaluation_result": "缺少评测结果",
    "batch_case_failed": "单条 Case 执行失败",
}


def _load_xlsx_case_result(output_dir: Path, row: dict[str, Any]) -> dict[str, Any]:
    case_id = str(row.get("case_id") or "")
    result_path = output_dir / "runs" / case_id / "evaluation_result.json"
    if result_path.exists():
        try:
            payload = json.loads(result_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            payload = {}
    else:
        payload = {}
    payload.setdefault("case_id", case_id)
    payload.setdefault("video_id", row.get("video_id"))
    payload.setdefault("instruction", row.get("instruction"))
    payload["_xlsx_result_path"] = str(result_path)
    return payload


def _na(value: Any, *, placeholder: str = "不适用") -> Any:
    return placeholder if value is None or value == "" else value


def _dash(value: Any) -> Any:
    return "—" if value is None or value == "" else value


def _yes_no(value: Any) -> str:
    return "是" if bool(value) else "否"


def _as_percent(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return float(value)
    return "不适用"


def _score_value(value: Any) -> Any:
    return round(float(value), 2) if isinstance(value, (int, float)) else "不适用"


def _token_or_dash(value: Any) -> Any:
    return value if isinstance(value, (int, float)) else "—"


def _result_case_id(result: dict[str, Any]) -> str:
    return str(result.get("case_id") or "")


def _status_label(status: Any) -> str:
    text = str(status or "")
    return STATUS_LABELS.get(text, text or "不适用")


def _tested_capability_label(result: dict[str, Any]) -> str:
    metadata = result.get("case_metadata") if isinstance(result.get("case_metadata"), dict) else {}
    tested = str(metadata.get("tested_capability") or "").strip()
    if tested:
        return tested
    primary = str(metadata.get("primary_capability") or "").strip()
    return USER_REQUIREMENT_LABELS.get(primary, primary or "不适用")


def _issue_text(result: dict[str, Any]) -> str:
    issues = result.get("issue_summary") if isinstance(result.get("issue_summary"), list) else []
    texts = []
    for issue in issues:
        if not isinstance(issue, dict):
            continue
        text = issue.get("display_text") or issue.get("label") or issue.get("issue_type")
        if text:
            texts.append(str(text))
    return "；".join(texts) if texts else "未发现明显问题"


def _judge_result(result: dict[str, Any]) -> dict[str, Any]:
    aesthetic = result.get("aesthetic_judge") if isinstance(result.get("aesthetic_judge"), dict) else {}
    judge_results = aesthetic.get("judge_results")
    if isinstance(judge_results, list) and judge_results and isinstance(judge_results[0], dict):
        return judge_results[0]
    return {}


def _judge_dimension(result: dict[str, Any], key: str) -> Any:
    judge = _judge_result(result)
    details = judge.get("dimension_details") if isinstance(judge.get("dimension_details"), dict) else {}
    detail = details.get(key) if isinstance(details.get(key), dict) else {}
    if detail.get("score") is not None:
        return detail.get("score")
    scores = judge.get("scores") if isinstance(judge.get("scores"), dict) else {}
    return _na(scores.get(key))


def _judge_issues(result: dict[str, Any]) -> str:
    judge = _judge_result(result)
    issues = judge.get("issues")
    if isinstance(issues, list) and issues:
        return "；".join(str(item.get("description") or item.get("label") or item.get("issue_type") or item) for item in issues)
    summary = judge.get("judge_summary")
    return str(summary) if summary else "不适用"


def _xlsx_all_case_rows(output_dir: Path, rows: list[dict[str, Any]], results: list[dict[str, Any]]) -> dict[str, Any]:
    headers = [
        "Case ID",
        "视频 ID",
        "用户指令",
        "测试内容",
        "执行状态",
        "内容选择得分",
        "成片观看体验得分",
        "综合得分",
        "是否触发兜底",
        "是否需要人工核查",
        "主要问题",
        "详细结果路径",
    ]
    records = []
    for row, result in zip(rows, results):
        records.append({
            "Case ID": _result_case_id(result),
            "视频 ID": result.get("video_id") or row.get("video_id"),
            "用户指令": result.get("instruction") or row.get("instruction"),
            "测试内容": _tested_capability_label(result),
            "执行状态": _status_label(result.get("evaluation_status")),
            "内容选择得分": _score_value(result.get("selection_score_v1")),
            "成片观看体验得分": _score_value(result.get("editing_experience_score_v1")),
            "综合得分": _score_value(result.get("final_score_v2")),
            "是否触发兜底": _yes_no((result.get("fallback_info") or {}).get("fallback_used")),
            "是否需要人工核查": _yes_no((result.get("manual_review") or {}).get("required")),
            "主要问题": _issue_text(result),
            "详细结果路径": str(output_dir / "runs" / _result_case_id(result) / "evaluation_result.json"),
        })
    return {"headers": headers, "rows": records, "score_columns": ["内容选择得分", "成片观看体验得分", "综合得分"]}


def _xlsx_selection_rows(results: list[dict[str, Any]]) -> dict[str, Any]:
    headers = [
        "Case ID",
        "指令解析模式",
        "内容筛选范围",
        "目标内容覆盖率",
        "合理内容占比",
        "默认高光占比",
        "明确禁止内容混入比例",
        "默认规避内容混入比例",
        "时长约束得分",
        "内容选择得分",
    ]
    records = []
    for result in results:
        metrics = result.get("time_metrics") if isinstance(result.get("time_metrics"), dict) else {}
        duration = result.get("duration_context") if isinstance(result.get("duration_context"), dict) else {}
        records.append({
            "Case ID": _result_case_id(result),
            "指令解析模式": _na(result.get("instruction_mode")),
            "内容筛选范围": _na(result.get("selection_scope")),
            "目标内容覆盖率": _as_percent(metrics.get("relevant_duration_coverage")),
            "合理内容占比": _as_percent(metrics.get("acceptable_precision")),
            "默认高光占比": _as_percent(metrics.get("default_highlight_precision")),
            "明确禁止内容混入比例": _as_percent(metrics.get("forbidden_duration_ratio")),
            "默认规避内容混入比例": _as_percent(metrics.get("avoid_by_default_overlap_ratio")),
            "时长约束得分": _score_value(duration.get("duration_score") if duration.get("duration_score") is not None else metrics.get("duration_score")),
            "内容选择得分": _score_value(result.get("selection_score_v1")),
        })
    return {
        "headers": headers,
        "rows": records,
        "percent_columns": ["目标内容覆盖率", "合理内容占比", "默认高光占比", "明确禁止内容混入比例", "默认规避内容混入比例"],
        "score_columns": ["时长约束得分", "内容选择得分"],
    }


def _xlsx_editing_rows(results: list[dict[str, Any]]) -> dict[str, Any]:
    headers = [
        "Case ID",
        "动作或口播边界完整性",
        "镜头衔接连贯性",
        "节奏紧凑性",
        "音画连续性",
        "独立可观看性",
        "成片观看体验得分",
        "Judge 识别的问题",
        "是否建议人工核查",
    ]
    records = []
    for result in results:
        records.append({
            "Case ID": _result_case_id(result),
            "动作或口播边界完整性": _judge_dimension(result, "clip_boundary_completeness"),
            "镜头衔接连贯性": _judge_dimension(result, "transition_coherence"),
            "节奏紧凑性": _judge_dimension(result, "pacing_and_conciseness"),
            "音画连续性": _judge_dimension(result, "audio_visual_continuity"),
            "独立可观看性": _judge_dimension(result, "standalone_watchability"),
            "成片观看体验得分": _score_value(result.get("editing_experience_score_v1")),
            "Judge 识别的问题": _judge_issues(result),
            "是否建议人工核查": _yes_no((result.get("manual_review") or {}).get("recommended")),
        })
    return {"headers": headers, "rows": records, "score_columns": headers[1:7]}


def _xlsx_technical_rows(results: list[dict[str, Any]]) -> dict[str, Any]:
    headers = [
        "Case ID",
        "是否可完整解码",
        "成片时长",
        "成片时长误差比例",
        "黑屏比例",
        "静帧比例",
        "静音比例",
        "重复片段比例",
        "技术质量是否通过",
        "技术问题说明",
    ]
    records = []
    for result in results:
        technical = result.get("technical_quality") if isinstance(result.get("technical_quality"), dict) else {}
        issues = (technical.get("technical_quality_errors") or []) + (technical.get("technical_quality_warnings") or [])
        records.append({
            "Case ID": _result_case_id(result),
            "是否可完整解码": _na(_yes_no(technical.get("decode_success")) if technical.get("decode_success") is not None else None),
            "成片时长": _na(technical.get("rendered_duration")),
            "成片时长误差比例": _as_percent(technical.get("rendered_duration_error_ratio")),
            "黑屏比例": _as_percent(technical.get("black_frame_ratio")),
            "静帧比例": _as_percent(technical.get("freeze_frame_ratio")),
            "静音比例": _as_percent(technical.get("silence_ratio")),
            "重复片段比例": _as_percent(technical.get("duplicate_source_ratio")),
            "技术质量是否通过": _na(_yes_no(technical.get("technical_quality_passed")) if technical.get("technical_quality_passed") is not None else None),
            "技术问题说明": "；".join(str(item) for item in issues) if issues else "不适用",
        })
    return {
        "headers": headers,
        "rows": records,
        "percent_columns": ["成片时长误差比例", "黑屏比例", "静帧比例", "静音比例", "重复片段比例"],
        "score_columns": ["成片时长"],
    }


def _xlsx_consumption_rows(results: list[dict[str, Any]]) -> dict[str, Any]:
    headers = [
        "Case ID",
        "视频剪辑总耗时",
        "预览生成耗时",
        "Skill Ark 调用耗时",
        "ffmpeg 渲染耗时",
        "Skill Ark Prompt Tokens",
        "Skill Ark Completion Tokens",
        "Skill Ark 总 Tokens",
        "自动评测总耗时",
        "Resolver 调用耗时",
        "Resolver 总 Tokens",
        "Judge 调用耗时",
        "Judge 总 Tokens",
        "完整链路总耗时",
        "完整链路总 Tokens",
    ]
    records = []
    for result in results:
        consumption = result.get("consumption") if isinstance(result.get("consumption"), dict) else {}
        editing = consumption.get("video_editing") if isinstance(consumption.get("video_editing"), dict) else {}
        evaluation = consumption.get("evaluation") if isinstance(consumption.get("evaluation"), dict) else {}
        end_to_end = consumption.get("end_to_end") if isinstance(consumption.get("end_to_end"), dict) else {}
        records.append({
            "Case ID": _result_case_id(result),
            "视频剪辑总耗时": _dash(editing.get("elapsed_seconds")),
            "预览生成耗时": _dash(editing.get("preview_generation_seconds")),
            "Skill Ark 调用耗时": _dash(editing.get("skill_llm_latency_seconds")),
            "ffmpeg 渲染耗时": _dash(editing.get("ffmpeg_render_seconds")),
            "Skill Ark Prompt Tokens": _token_or_dash(editing.get("llm_prompt_tokens")),
            "Skill Ark Completion Tokens": _token_or_dash(editing.get("llm_completion_tokens")),
            "Skill Ark 总 Tokens": _token_or_dash(editing.get("llm_total_tokens")),
            "自动评测总耗时": _dash(evaluation.get("elapsed_seconds")),
            "Resolver 调用耗时": _dash(evaluation.get("resolver_latency_seconds")),
            "Resolver 总 Tokens": _token_or_dash(evaluation.get("resolver_total_tokens")),
            "Judge 调用耗时": _dash(evaluation.get("judge_latency_seconds")),
            "Judge 总 Tokens": _token_or_dash(evaluation.get("judge_total_tokens")),
            "完整链路总耗时": _dash(end_to_end.get("elapsed_seconds")),
            "完整链路总 Tokens": _token_or_dash(end_to_end.get("llm_total_tokens")),
        })
    return {"headers": headers, "rows": records, "score_columns": ["视频剪辑总耗时", "预览生成耗时", "Skill Ark 调用耗时", "ffmpeg 渲染耗时", "自动评测总耗时", "Resolver 调用耗时", "Judge 调用耗时", "完整链路总耗时"]}


def _xlsx_metric_description_rows() -> dict[str, Any]:
    rows = [
        {"指标名称": "内容选择得分", "说明": "衡量成片片段是否覆盖应选内容，并避免不相关或明确禁止内容。"},
        {"指标名称": "成片观看体验得分", "说明": "由自动 Judge 评估动作/口播边界、镜头衔接、节奏、音画连续性和独立可观看性。"},
        {"指标名称": "综合得分", "说明": "综合得分 = 70% 内容选择得分 + 30% 成片观看体验得分。"},
        {"指标名称": "正式计分资格", "说明": "只有完整评分、Ark 正式输出、产物校验通过且未被排除的 Case 进入正式平均分。"},
        {"指标名称": "目标内容覆盖率", "说明": "预测片段覆盖目标相关内容的比例。"},
        {"指标名称": "合理内容占比", "说明": "预测片段中 relevant 与 allowed_context 的合计占比。"},
        {"指标名称": "默认高光占比", "说明": "通用指令下预测内容命中默认高光的比例。"},
        {"指标名称": "明确禁止内容混入比例", "说明": "预测片段中混入用户明确禁止内容的比例。"},
        {"指标名称": "默认规避内容混入比例", "说明": "预测片段中混入 GT 标记 avoid_by_default 内容的比例。"},
        {"指标名称": "时长约束得分", "说明": "成片时长满足用户或系统时长约束的程度。"},
        {"指标名称": "技术质量是否通过", "说明": "检查解码、黑屏、静帧、静音、重复片段和时长一致性等硬性质量。"},
        {"指标名称": "是否触发兜底", "说明": "表示 Skill Ark 调用失败后是否进入 fallback/mock 诊断路径。"},
        {"指标名称": "是否需要人工核查", "说明": "表示自动评测认为该 Case 需要人工确认结果。"},
    ]
    return {"headers": ["指标名称", "说明"], "rows": rows}


def _xlsx_column_width(header: str, worksheet: Any, index: int) -> float:
    max_length = len(str(header))
    for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2):
        for item in cell:
            max_length = max(max_length, len(str(item.value or "")))
    if header in {"用户指令", "测试内容", "主要问题", "详细结果路径", "Judge 识别的问题", "技术问题说明", "说明"}:
        return min(max(max_length + 2, 24), 60)
    return min(max(max_length + 2, 12), 28)


def _display_text(value: Any, *, placeholder: str = "—") -> str:
    if value is None or value == "":
        return placeholder
    if isinstance(value, bool):
        return "是" if value else "否"
    return str(value)


def _format_score_for_html(value: Any) -> str:
    return f"{float(value):.2f}" if isinstance(value, (int, float)) else "—"


def _format_seconds_for_html(value: Any) -> str:
    return f"{float(value):.2f} 秒" if isinstance(value, (int, float)) else "—"


def _format_percent_for_html(value: Any) -> str:
    return f"{float(value) * 100:.2f}%" if isinstance(value, (int, float)) else "不适用"


def _format_basis_value(value: Any, *, percent: bool = False, score: bool = False) -> str:
    if percent:
        return _format_percent_for_html(value)
    if score:
        return _format_score_for_html(value) if isinstance(value, (int, float)) else "不适用"
    return _display_text(value, placeholder="不适用")


def _case_metadata(result: dict[str, Any]) -> dict[str, Any]:
    return result.get("case_metadata") if isinstance(result.get("case_metadata"), dict) else {}


def _scenario_labels(result: dict[str, Any]) -> str:
    metadata = _case_metadata(result)
    raw = metadata.get("challenge_tags")
    if isinstance(raw, list):
        values = raw
    elif isinstance(raw, str):
        values = raw.split(",")
    else:
        values = []
    labels = []
    for value in values:
        key = str(value).strip()
        if not key or key in IGNORED_CHALLENGE_TAGS:
            continue
        labels.append(VIDEO_SCENARIO_LABELS.get(key, key))
    return "、".join(labels) if labels else "—"


def _technical_quality_note(result: dict[str, Any]) -> str:
    technical = result.get("technical_quality") if isinstance(result.get("technical_quality"), dict) else {}
    issues = (technical.get("technical_quality_errors") or []) + (technical.get("technical_quality_warnings") or [])
    if issues:
        return "；".join(str(item) for item in issues)
    if technical.get("technical_quality_passed") is True:
        return "技术质量检查通过"
    return "不适用"


def _failure_reason(result: dict[str, Any]) -> str:
    failure = result.get("failure_info") if isinstance(result.get("failure_info"), dict) else {}
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    reasons = []
    for value in [
        failure.get("error_message"),
        result.get("error_message"),
    ]:
        if value:
            reasons.append(str(value))
    for value in artifact.get("artifact_validation_errors") or []:
        if value:
            reasons.append(str(value))
    return "；".join(reasons) if reasons else "—"


def _manual_review_status(result: dict[str, Any]) -> tuple[str, str]:
    review = result.get("manual_review") if isinstance(result.get("manual_review"), dict) else {}
    if review.get("required"):
        return "必须人工核查", _issue_text(result)
    if review.get("recommended"):
        return "建议人工抽查", _issue_text(result)
    return "无需人工核查", "无需人工核查"


def _read_json_file(path: Any) -> dict[str, Any]:
    if not path:
        return {}
    candidate = Path(str(path))
    if not candidate.exists() or not candidate.is_file():
        return {}
    try:
        payload = json.loads(candidate.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _artifact_path(result: dict[str, Any], key: str) -> str:
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    paths = artifact.get("paths") if isinstance(artifact.get("paths"), dict) else {}
    value = paths.get(key)
    return str(value or "")


def _selected_segments(result: dict[str, Any]) -> list[dict[str, Any]]:
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    summary = artifact.get("result_summary") if isinstance(artifact.get("result_summary"), dict) else {}
    if isinstance(summary.get("final_segments"), list):
        return [item for item in summary["final_segments"] if isinstance(item, dict)]
    segments_path = _artifact_path(result, "segments_json") or summary.get("segments_json")
    segments_payload = _read_json_file(segments_path)
    segments = segments_payload.get("final_segments")
    return [item for item in segments if isinstance(item, dict)] if isinstance(segments, list) else []


def _relative_file_link(label: str, target: Any, from_dir: Path) -> str:
    if not target:
        return f"<span class=\"missing\">{html.escape(label)}：文件不存在</span>"
    target_path = Path(str(target))
    if not target_path.exists():
        return f"<span class=\"missing\">{html.escape(label)}：文件不存在</span>"
    href = os.path.relpath(target_path, start=from_dir)
    return f"<a href=\"{html.escape(href)}\">{html.escape(label)}</a>"


def _html_table(rows: list[tuple[str, str]]) -> str:
    body = "".join(
        f"<tr><th>{html.escape(label)}</th><td>{html.escape(value)}</td></tr>"
        for label, value in rows
    )
    return f"<table class=\"kv\"><tbody>{body}</tbody></table>"


def _score_cards_html(result: dict[str, Any]) -> str:
    cards = [
        ("内容选择得分", result.get("selection_score_v1")),
        ("成片观看体验得分", result.get("editing_experience_score_v1")),
        ("综合得分", result.get("final_score_v2")),
    ]
    return "".join(
        f"<div class=\"score-card\"><span>{html.escape(label)}</span><strong>{html.escape(_format_score_for_html(value))}</strong></div>"
        for label, value in cards
    )


def _basis_rows(result: dict[str, Any]) -> list[tuple[str, str]]:
    metrics = result.get("time_metrics") if isinstance(result.get("time_metrics"), dict) else {}
    duration = result.get("duration_context") if isinstance(result.get("duration_context"), dict) else {}
    duration_score = duration.get("duration_score") if duration.get("duration_score") is not None else metrics.get("duration_score")
    return [
        ("目标内容覆盖率", _format_basis_value(metrics.get("relevant_duration_coverage"), percent=True)),
        ("合理内容占比", _format_basis_value(metrics.get("acceptable_precision"), percent=True)),
        ("默认高光占比", _format_basis_value(metrics.get("default_highlight_precision"), percent=True)),
        ("明确禁止内容混入比例", _format_basis_value(metrics.get("forbidden_duration_ratio"), percent=True)),
        ("默认规避内容混入比例", _format_basis_value(metrics.get("avoid_by_default_overlap_ratio"), percent=True)),
        ("时长约束得分", _format_basis_value(duration_score, score=True)),
        ("技术质量说明", _technical_quality_note(result)),
    ]


def _segments_table_html(result: dict[str, Any]) -> str:
    segments = _selected_segments(result)
    if not segments:
        return "<div class=\"empty\">未找到可展示的片段信息。</div>"
    rows = []
    for index, segment in enumerate(segments, start=1):
        start = segment.get("start")
        end = segment.get("end")
        duration = None
        if isinstance(start, (int, float)) and isinstance(end, (int, float)):
            duration = max(0.0, float(end) - float(start))
        rows.append(
            "<tr>"
            f"<td>{index}</td>"
            f"<td>{html.escape(_format_seconds_for_html(start))}</td>"
            f"<td>{html.escape(_format_seconds_for_html(end))}</td>"
            f"<td>{html.escape(_format_seconds_for_html(duration))}</td>"
            f"<td>{html.escape(_display_text(segment.get('title')))}</td>"
            "</tr>"
        )
    return (
        "<table><thead><tr><th>序号</th><th>开始时间</th><th>结束时间</th><th>片段时长</th><th>片段标题</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table>"
    )


def _consumption_rows(result: dict[str, Any]) -> list[tuple[str, str]]:
    consumption = result.get("consumption") if isinstance(result.get("consumption"), dict) else {}
    editing = consumption.get("video_editing") if isinstance(consumption.get("video_editing"), dict) else {}
    evaluation = consumption.get("evaluation") if isinstance(consumption.get("evaluation"), dict) else {}
    end_to_end = consumption.get("end_to_end") if isinstance(consumption.get("end_to_end"), dict) else {}
    return [
        ("视频剪辑总耗时", _format_seconds_for_html(editing.get("elapsed_seconds"))),
        ("预览生成耗时", _format_seconds_for_html(editing.get("preview_generation_seconds"))),
        ("Skill Ark 调用耗时", _format_seconds_for_html(editing.get("skill_llm_latency_seconds"))),
        ("ffmpeg 渲染耗时", _format_seconds_for_html(editing.get("ffmpeg_render_seconds"))),
        ("Skill Ark 总 Tokens", _display_text(editing.get("llm_total_tokens"))),
        ("自动评测总耗时", _format_seconds_for_html(evaluation.get("elapsed_seconds"))),
        ("Resolver 调用耗时", _format_seconds_for_html(evaluation.get("resolver_latency_seconds"))),
        ("Resolver 总 Tokens", _display_text(evaluation.get("resolver_total_tokens"))),
        ("Judge 调用耗时", _format_seconds_for_html(evaluation.get("judge_latency_seconds"))),
        ("Judge 总 Tokens", _display_text(evaluation.get("judge_total_tokens"))),
        ("完整链路总耗时", _format_seconds_for_html(end_to_end.get("elapsed_seconds"))),
        ("完整链路总 Tokens", _display_text(end_to_end.get("llm_total_tokens"))),
    ]


def _file_entry_rows(output_dir: Path, cases_dir: Path, result: dict[str, Any]) -> list[tuple[str, str]]:
    case_id = _result_case_id(result)
    run_dir = output_dir / "runs" / case_id
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    paths = artifact.get("paths") if isinstance(artifact.get("paths"), dict) else {}
    summary = artifact.get("result_summary") if isinstance(artifact.get("result_summary"), dict) else {}
    rows = [
        ("查看原始评测结果", run_dir / "evaluation_result.json"),
        ("查看冻结评分规则", run_dir / "generated_case.json"),
        ("查看运行元信息", run_dir / "run_manifest.json"),
        ("查看 Skill 结果摘要", paths.get("result_summary")),
        ("查看运行日志", paths.get("run_log") or summary.get("run_log")),
        ("查看最终成片", paths.get("highlight_video") or summary.get("highlight_video")),
        ("返回总览报告", output_dir / "report.html"),
    ]
    return [(label, _relative_file_link(label, target, cases_dir)) for label, target in rows]


def _case_detail_html(output_dir: Path, cases_dir: Path, row: dict[str, Any], result: dict[str, Any]) -> str:
    case_id = _result_case_id(result) or str(row.get("case_id") or "")
    metadata = _case_metadata(result)
    artifact = result.get("artifact_validation") if isinstance(result.get("artifact_validation"), dict) else {}
    official = result.get("official_score_eligibility") if isinstance(result.get("official_score_eligibility"), dict) else {}
    fallback = result.get("fallback_info") if isinstance(result.get("fallback_info"), dict) else {}
    review = result.get("manual_review") if isinstance(result.get("manual_review"), dict) else {}
    technical = result.get("technical_quality") if isinstance(result.get("technical_quality"), dict) else {}
    status = _status_label(result.get("evaluation_status"))
    failure_reason = _failure_reason(result)
    manual_status, manual_reason = _manual_review_status(result)
    fallback_notice = ""
    if fallback.get("fallback_used"):
        fallback_notice = (
            "<div class=\"notice orange\">Ark 调用失败，已触发兜底结果。<br>"
            "该结果仅用于诊断，不进入正式平均分。</div>"
        )
    failure_notice = ""
    if failure_reason != "—":
        failure_notice = f"<div class=\"notice red\"><strong>失败原因</strong><br>{html.escape(failure_reason)}</div>"
    base_rows = [
        ("Case ID", case_id),
        ("视频 ID", _display_text(result.get("video_id") or row.get("video_id"))),
        ("用户指令", _display_text(result.get("instruction") or row.get("instruction"))),
        ("测试内容", _tested_capability_label(result)),
        ("视频场景标签", _scenario_labels(result)),
    ]
    status_rows = [
        ("执行状态", status),
        ("是否进入正式计分", _display_text(bool(official.get("eligible")))),
        ("是否触发兜底", _display_text(bool(fallback.get("fallback_used")))),
        ("是否需要人工核查", _display_text(bool(review.get("required")))),
        ("技术质量是否通过", _display_text(technical.get("technical_quality_passed"))),
    ]
    issue_rows = [
        ("主要问题", _issue_text(result)),
        ("人工核查状态", manual_status),
        ("人工核查原因", manual_reason),
    ]
    file_rows = _file_entry_rows(output_dir, cases_dir, result)
    file_table = "".join(
        f"<tr><th>{html.escape(label)}</th><td>{value}</td></tr>"
        for label, value in file_rows
    )
    return f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><title>{html.escape(case_id)} - Case 详情</title><style>
:root{{--blue:#12365f;--bg:#f5f7fb;--text:#1f2937;--muted:#667085;--line:#d9e1ec;--green:#18864b;--orange:#b65f00;--red:#b42318;--info:#175cd3}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',sans-serif;line-height:1.55}}.page{{max-width:1200px;margin:0 auto;padding:32px 28px 56px}}h1{{margin:0;color:var(--blue);font-size:30px}}h2{{margin:0 0 14px;color:var(--blue);font-size:21px}}section{{margin-top:20px}}.card{{background:#fff;border:1px solid var(--line);border-radius:14px;padding:20px;box-shadow:0 8px 24px rgba(18,54,95,.05)}}table{{width:100%;border-collapse:separate;border-spacing:0;border:1px solid var(--line);border-radius:12px;overflow:hidden;background:#fff}}th,td{{padding:11px 12px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}}th{{width:220px;background:#f0f4f9;color:#344054;font-weight:650}}tr:last-child th,tr:last-child td{{border-bottom:none}}.score-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:14px}}.score-card{{background:#fff;border:1px solid var(--line);border-radius:14px;padding:18px}}.score-card span{{display:block;color:var(--muted);font-size:14px}}.score-card strong{{display:block;margin-top:8px;color:var(--blue);font-size:32px}}.note{{margin-top:10px;color:var(--muted);font-size:14px}}.notice{{border-radius:12px;padding:14px 16px;margin-top:12px}}.orange{{background:#fff4e5;border:1px solid #ffd39a;color:var(--orange)}}.red{{background:#fde9e7;border:1px solid #ffb5ae;color:var(--red)}}.empty{{background:#f1f3f6;border:1px dashed #b8c2d0;border-radius:12px;padding:16px;color:#475467}}.missing{{color:#98a2b3}}a{{color:#175cd3;text-decoration:none}}a:hover{{text-decoration:underline}}.top-link{{margin-top:8px}}.top-link a{{font-weight:650}}
</style></head><body><main class="page">
<section class="card"><h1>{html.escape(case_id)} Case 详情</h1><div class="top-link"><a href="../report.html">返回总览报告</a></div></section>
<section class="card"><h2>Case 基本信息</h2>{_html_table(base_rows)}</section>
<section class="card"><h2>执行状态</h2>{_html_table(status_rows)}{fallback_notice}{failure_notice}</section>
<section><h2>核心得分</h2><div class="score-grid">{_score_cards_html(result)}</div><p class="note">综合得分 = 70% 内容选择得分 + 30% 成片观看体验得分</p></section>
<section class="card"><h2>得分依据摘要</h2>{_html_table(_basis_rows(result))}</section>
<section class="card"><h2>实际选中的片段</h2>{_segments_table_html(result)}</section>
<section class="card"><h2>主要问题与人工核查</h2>{_html_table(issue_rows)}</section>
<section class="card"><h2>耗时与 Token</h2>{_html_table(_consumption_rows(result))}</section>
<section class="card"><h2>原始文件入口</h2><table class="kv"><tbody>{file_table}</tbody></table></section>
</main></body></html>"""


def _select_columns(rows: list[dict[str, Any]], fields: list[str]) -> list[dict[str, Any]]:
    return [{field: row.get(field) for field in fields} for row in rows]


def _write_minimal_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _xlsx_content_types(len(sheets)))
        archive.writestr("_rels/.rels", _xlsx_root_rels())
        archive.writestr("xl/workbook.xml", _xlsx_workbook(list(sheets)))
        archive.writestr("xl/_rels/workbook.xml.rels", _xlsx_workbook_rels(len(sheets)))
        for index, (_, rows) in enumerate(sheets.items(), start=1):
            archive.writestr(f"xl/worksheets/sheet{index}.xml", _xlsx_sheet(rows))


def _xlsx_content_types(sheet_count: int) -> str:
    overrides = "".join(
        f'<Override PartName="/xl/worksheets/sheet{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        for i in range(1, sheet_count + 1)
    )
    return f'<?xml version="1.0" encoding="UTF-8"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>{overrides}</Types>'


def _xlsx_root_rels() -> str:
    return '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'


def _xlsx_workbook(names: list[str]) -> str:
    sheets = "".join(
        f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>'
        for index, name in enumerate(names, start=1)
    )
    return f'<?xml version="1.0" encoding="UTF-8"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets>{sheets}</sheets></workbook>'


def _xlsx_workbook_rels(sheet_count: int) -> str:
    relationships = "".join(
        f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i}.xml"/>'
        for i in range(1, sheet_count + 1)
    )
    return f'<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{relationships}</Relationships>'


def _xlsx_sheet(rows: list[dict[str, Any]]) -> str:
    fields = list(rows[0].keys()) if rows else ["case_id"]
    all_rows = [fields] + [[row.get(field) for field in fields] for row in rows]
    xml_rows = []
    for row_index, values in enumerate(all_rows, start=1):
        cells = []
        for col_index, value in enumerate(values, start=1):
            ref = f"{_col_name(col_index)}{row_index}"
            text = "" if value is None else str(value)
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{escape(text)}</t></is></c>')
        xml_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return f'<?xml version="1.0" encoding="UTF-8"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>{"".join(xml_rows)}</sheetData></worksheet>'


def _col_name(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _dashboard_payload(summary: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
    case_ids = [str(row.get("case_id") or "") for row in rows]
    failure_ids = list(summary.get("failure_summary", {}).get("case_ids", []))
    fallback_ids = list(summary.get("fallback_summary", {}).get("case_ids", []))
    manual = summary.get("manual_review_summary", {})
    generated_ids = [
        str(row.get("case_id") or "")
        for row in rows
        if row.get("evaluation_status") not in {"invalid_artifact", "missing_evaluation_result"}
    ]
    scored_ids = [
        str(row.get("case_id") or "")
        for row in rows
        if row.get("official_eligible") is True
    ]
    technical_values = [
        row.get("technical_quality_passed")
        for row in rows
        if row.get("technical_quality_passed") in {True, False}
    ]
    technical_pass_rate = None
    if technical_values:
        technical_pass_rate = f"{sum(1 for value in technical_values if value is True) / len(technical_values) * 100:.2f}%"
    overall = summary.get("overall_scores", {})
    execution = summary.get("execution_overview", {})
    special = summary.get("special_reports_overview") or {}
    return {
        "info": {
            "case_count": summary.get("run_metadata", {}).get("case_count"),
            "generated_at": summary.get("run_metadata", {}).get("generated_at"),
        },
        "execution": {
            "planned": execution.get("planned_case_count"),
            "generated": execution.get("generated_result_count"),
            "scored": execution.get("official_scored_count"),
            "fallback": execution.get("fallback_diagnostic_count"),
            "failed": execution.get("failed_or_incomplete_count"),
            "manual_required": execution.get("manual_review_blocked_count"),
            "manual_recommended": manual.get("recommended_count"),
        },
        "scores": {
            "content": overall.get("selection_score_v1_avg"),
            "editing": overall.get("editing_experience_score_v1_avg"),
            "final": overall.get("final_score_v2_avg"),
            "technical_pass_rate": technical_pass_rate,
        },
        "groups": {
            "all": case_ids,
            "generated": generated_ids,
            "scored": scored_ids,
            "fallback": fallback_ids,
            "failed": failure_ids,
            "manual_required": list(manual.get("required_case_ids", [])),
            "manual_recommended": list(manual.get("recommended_case_ids", [])),
        },
        "user_requirements": summary.get("breakdowns", {}).get("by_user_requirement", []),
        "video_scenarios": summary.get("breakdowns", {}).get("by_video_scenario", []),
        "issues": summary.get("breakdowns", {}).get("by_issue", []),
        "special": (
            [
                {"label": str(key), "value": value}
                for key, value in special.items()
            ]
            if isinstance(special, dict) and special
            else []
        ),
        "cases": rows,
    }


def write_html_reports(output_dir: Path, summary: dict[str, Any], rows: list[dict[str, Any]], results: list[dict[str, Any]]) -> None:
    cases_dir = output_dir / "cases"
    cases_dir.mkdir(parents=True, exist_ok=True)
    dashboard = _dashboard_payload(summary, rows)
    data_json = json.dumps(dashboard, ensure_ascii=False)
    html_doc = f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><title>ClawCut V2 正式评测报告</title><style>
:root{{color-scheme:light;--blue:#12365f;--text:#1f2937;--muted:#667085;--bg:#f5f7fb;--line:#d9e1ec;--green:#18864b;--orange:#b65f00;--red:#b42318;--info:#175cd3}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',sans-serif;line-height:1.5}}.page{{max-width:1400px;margin:0 auto;padding:32px 28px 56px}}h1{{margin:0 0 8px;color:var(--blue);font-size:32px}}h2{{margin:0 0 16px;color:var(--blue);font-size:22px}}.subtle{{color:var(--muted)}}section{{margin-top:24px}}.panel{{background:#fff;border:1px solid var(--line);border-radius:14px;padding:20px;box-shadow:0 8px 24px rgba(18,54,95,.05)}}.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:14px}}.card{{appearance:none;border:1px solid var(--line);border-radius:14px;background:#fff;padding:16px;text-align:left;cursor:pointer;box-shadow:0 8px 20px rgba(18,54,95,.05)}}.card:hover,.click-row:hover{{border-color:#8fb5e8;background:#f8fbff}}.card-label{{font-size:14px;color:var(--muted)}}.card-value{{display:block;margin-top:8px;font-size:30px;font-weight:750;color:var(--blue)}}.green .card-value{{color:var(--green)}}.orange .card-value{{color:var(--orange)}}.red .card-value{{color:var(--red)}}.info .card-value{{color:var(--info)}}table{{width:100%;border-collapse:separate;border-spacing:0;border:1px solid var(--line);border-radius:12px;overflow:hidden;background:#fff}}th,td{{padding:11px 12px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}}th{{background:#f0f4f9;color:#344054;font-weight:650}}tr:last-child td{{border-bottom:none}}.click-row{{cursor:pointer}}.note{{margin-top:10px;color:var(--muted);font-size:14px}}.empty{{background:#f1f3f6;border:1px dashed #b8c2d0;border-radius:12px;padding:18px;color:#475467}}#drawer{{position:fixed;right:0;top:0;bottom:0;width:560px;max-width:94vw;background:#fff;border-left:1px solid var(--line);box-shadow:-12px 0 30px rgba(0,0,0,.12);padding:20px;overflow:auto;display:none;z-index:10}}.drawer-head{{display:flex;gap:12px;align-items:center;justify-content:space-between;margin-bottom:12px}}.close{{border:1px solid var(--line);background:#fff;border-radius:8px;padding:7px 10px;cursor:pointer}}.search{{width:100%;border:1px solid var(--line);border-radius:10px;padding:10px 12px;margin:8px 0 14px}}.case-card{{border:1px solid var(--line);border-radius:12px;padding:12px;margin:10px 0;background:#fbfcff}}.case-title{{font-weight:700;color:var(--blue)}}.case-meta{{display:grid;grid-template-columns:120px 1fr;gap:5px 10px;margin-top:8px;font-size:14px}}.pill{{display:inline-block;border-radius:999px;background:#eef4ff;color:#175cd3;padding:2px 8px;font-size:12px}}a{{color:#175cd3;text-decoration:none}}a:hover{{text-decoration:underline}}
</style></head><body><main class="page">
<section class="panel"><h1>ClawCut V2 正式评测报告</h1><div class="subtle">评测 Case 数：{dashboard['info']['case_count']}　生成时间：{html.escape(str(dashboard['info']['generated_at']))}</div></section>
<script type="application/json" id="report-data">{html.escape(data_json)}</script>
<section><h2>总体执行情况</h2><div class="grid" id="execution-cards"></div></section>
<section><h2>整体剪辑效果</h2><div class="grid" id="score-cards"></div><p class="note">综合得分 = 70% 内容选择得分 + 30% 成片观看体验得分</p></section>
<section class="panel"><h2>不同用户要求下的剪辑表现</h2><div id="user-table"></div></section>
<section class="panel"><h2>不同视频场景下的剪辑表现</h2><div id="scenario-table"></div><p class="note">同一条 Case 可以同时属于多个视频场景，因此各行数量之和可能大于总 Case 数。</p></section>
<section class="panel"><h2>本轮主要问题</h2><div id="issue-table"></div></section>
<section class="panel"><h2>专项测试概况</h2><div id="special-overview"></div></section>
</main><aside id="drawer"></aside><script>
const data = JSON.parse(document.getElementById('report-data').textContent);
const statusLabels = {{scored_complete:'正式计分完成', diagnostic_only:'仅用于诊断', invalid_artifact:'产物校验失败', judge_failed:'成片观看体验评测失败', judge_video_upload_failed:'Judge 视频上传失败', resolver_failed:'用户指令解析失败', missing_evaluation_result:'缺少评测结果', batch_case_failed:'单条 Case 执行失败'}};
const empty = v => v === null || v === undefined || v === '' ? '—' : v;
const score = v => typeof v === 'number' ? v.toFixed(2) : '—';
const casesById = new Map(data.cases.map(c => [c.case_id, c]));
function resolveCases(ids){{return ids.map(id=>casesById.get(id)).filter(Boolean);}}
function card(label,value,kind,ids){{return `<button class="card ${{kind||''}}" onclick='openDrawer(${{JSON.stringify(label)}}, ${{JSON.stringify(ids)}})'><span class="card-label">${{label}}</span><strong class="card-value">${{empty(value)}}</strong></button>`}}
function renderExecution(){{const e=data.execution;document.getElementById('execution-cards').innerHTML=[
card('计划评测 Case 数',e.planned,'',data.groups.all),
card('成功生成结果',e.generated,'',data.groups.generated),
card('进入正式计分',e.scored,'green',data.groups.scored),
card('触发兜底结果',e.fallback,'orange',data.groups.fallback),
card('执行失败或未完成',e.failed,'red',data.groups.failed),
card('必须人工核查',e.manual_required,'info',data.groups.manual_required),
card('建议人工抽查',e.manual_recommended,'info',data.groups.manual_recommended)
].join('');}}
function renderScores(){{const s=data.scores;document.getElementById('score-cards').innerHTML=[
card('平均内容选择得分',score(s.content),'',data.groups.scored),
card('平均成片观看体验得分',score(s.editing),'',data.groups.scored),
card('平均综合得分',score(s.final),'green',data.groups.scored),
card('成片技术质量通过率',s.technical_pass_rate,'',data.groups.generated)
].join('');}}
function table(headers,rows,kind){{if(!rows.length)return '<div class="empty">暂无数据。</div>';return `<table><thead><tr>${{headers.map(h=>`<th>${{h}}</th>`).join('')}}</tr></thead><tbody>${{rows.map(r=>`<tr class="click-row" onclick='openDrawer(${{JSON.stringify(r.label)}}, ${{JSON.stringify(r.case_ids||[])}})'>${{kind==='issue'?`<td>${{r.label}}</td><td>${{r.count}}</td>`:`<td>${{r.label}}</td><td>${{r.case_count}}</td><td>${{r.official_scored_count}}</td><td>${{score(r.average_content_selection_score)}}</td><td>${{score(r.average_editing_experience_score)}}</td><td>${{score(r.average_final_score)}}</td>`}}</tr>`).join('')}}</tbody></table>`}}
function renderTables(){{document.getElementById('user-table').innerHTML=table(['用户要求类型','Case 数','正式计分数','平均内容选择得分','平均成片观看体验得分','平均综合得分'],data.user_requirements,'normal');document.getElementById('scenario-table').innerHTML=table(['视频场景','Case 数','正式计分数','平均内容选择得分','平均成片观看体验得分','平均综合得分'],data.video_scenarios,'normal');document.getElementById('issue-table').innerHTML=table(['问题类型','涉及 Case 数'],data.issues,'issue');}}
function renderSpecial(){{const el=document.getElementById('special-overview');if(!data.special||!data.special.length){{el.innerHTML='<div class="empty"><strong>专项测试尚未执行。</strong><br>后续将补充评测规则解析、FPS 敏感性、运行稳定性和异常输入处理结果。</div>';return}}el.innerHTML='<table><tbody>'+data.special.map(x=>`<tr><td>${{x.label}}</td><td>${{x.value}}</td></tr>`).join('')+'</tbody></table>';}}
function openDrawer(title,ids){{const drawer=document.getElementById('drawer');drawer.style.display='block';drawer.dataset.ids=JSON.stringify(ids);drawer.innerHTML=`<div class="drawer-head"><h2>${{title}}</h2><button class="close" onclick="closeDrawer()">关闭</button></div><input class="search" id="case-search" placeholder="搜索 Case ID" oninput="renderDrawerCases()"><div id="drawer-list"></div>`;renderDrawerCases();}}
function closeDrawer(){{document.getElementById('drawer').style.display='none';}}
function renderDrawerCases(){{const drawer=document.getElementById('drawer');const ids=JSON.parse(drawer.dataset.ids||'[]');const q=(document.getElementById('case-search')?.value||'').trim().toLowerCase();const list=resolveCases(ids).filter(c=>!q||String(c.case_id).toLowerCase().includes(q));document.getElementById('drawer-list').innerHTML=list.length?list.map(c=>`<div class="case-card"><div class="case-title"><a href="cases/${{c.case_id}}.html">${{c.case_id}}</a></div><div class="case-meta"><span>视频 ID</span><span>${{empty(c.video_id)}}</span><span>用户指令</span><span>${{empty(c.instruction)}}</span><span>执行状态</span><span><span class="pill">${{statusLabels[c.evaluation_status]||empty(c.evaluation_status)}}</span></span><span>内容选择得分</span><span>${{score(c.selection_score_v1)}}</span><span>成片观看体验得分</span><span>${{score(c.editing_experience_score_v1)}}</span><span>综合得分</span><span>${{score(c.final_score_v2)}}</span><span>主要问题</span><span>${{empty(c.issues)}}</span></div></div>`).join(''):'<div class="empty">没有匹配的 Case。</div>';}}
renderExecution();renderScores();renderTables();renderSpecial();
</script></body></html>"""
    (output_dir / "report.html").write_text(html_doc, encoding="utf-8")
    for row, result in zip(rows, results):
        case_html = _case_detail_html(output_dir, cases_dir, row, result)
        (cases_dir / f"{row['case_id']}.html").write_text(case_html, encoding="utf-8")


def run(args: argparse.Namespace) -> dict[str, Any]:
    path_map = parse_path_maps(args.path_map)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    cases = read_jsonl(args.cases)
    results: list[dict[str, Any]] = []
    started = time.monotonic()
    for index, case in enumerate(cases, start=1):
        if args.mode == "report-only":
            result = load_existing_result(case, index, args.output_dir)
        else:
            result = run_case(case=case, index=index, args=args, path_map=path_map)
        results.append(result)
    rows = [flat_case_row(case, result) for case, result in zip(cases, results)]
    summary = build_summary(cases, results, rows, args.output_dir)
    summary["run_metadata"]["elapsed_seconds"] = round(time.monotonic() - started, 3)
    write_json(args.output_dir / "summary.json", summary)
    write_csv(args.output_dir / "case_results.csv", rows)
    write_xlsx(args.output_dir / "case_results.xlsx", rows)
    write_html_reports(args.output_dir, summary, rows, results)
    return summary


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run ClawCut official V2 evaluation and report generation.")
    parser.add_argument("--mode", choices=["evaluate", "report-only"], default="evaluate")
    parser.add_argument("--cases", type=Path, required=True)
    parser.add_argument("--gt-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--path-map", action="append", default=[])
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--retry-failed", action="store_true")
    parser.add_argument("--technical_quality_config", type=Path, default=Path("evaluation/config/default.yaml"))
    parser.add_argument("--auto_upload_judge_video", action="store_true")
    parser.add_argument("--tos_bucket", default=None)
    parser.add_argument("--tos_region", default=None)
    parser.add_argument("--tos_endpoint", default=None)
    parser.add_argument("--tos_key_prefix", default=None)
    parser.add_argument("--tos_presign_expires_seconds", type=int)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    summary = run(parse_args(argv))
    print(f"official eval report v2 完成：{summary['artifact_paths']['report_html']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
